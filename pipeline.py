#!/usr/bin/env python3
"""
Riddle pipeline: pick the next unposted riddle -> render (with variation) ->
add cinematic countdown audio -> emit a manifest for the uploader.

Posting order (launch strategy): ORIGINAL riddles first (owned IP, lowest policy
exposure), then the main bank. Weekly palette rotation is driven by post order.

State of truth = the workbook's "Posted?" column. This module does NOT mark a
riddle posted; the uploader calls mark_posted() only AFTER a successful upload,
so a failed run never burns a riddle.

CLI (local test):
    python3 pipeline.py            # build next video + audio, write manifest
    python3 pipeline.py --no-audio # skip audio (faster)
    python3 pipeline.py --peek     # just show what's next, render nothing
"""
import os, sys, json, subprocess, datetime, re
from openpyxl import load_workbook
import render_video
import metadata

# Where the scripts live (so we can find add_audio.py next to us).
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# Where data lives. Locally defaults to the workspace folder; on GitHub Actions
# set RIDDLE_ROOT=. (repo root). XLSX path can be overridden independently.
DATA_ROOT = os.environ.get("RIDDLE_ROOT",
                           "/sessions/magical-modest-ramanujan/mnt/Youtube")
XLSX  = os.environ.get("RIDDLE_XLSX", f"{DATA_ROOT}/Riddle_Content_Bank.xlsx")
OUTDIR = f"{DATA_ROOT}/Videos"
MANIFEST = f"{DATA_ROOT}/next_video.json"

# Ordered list of adult sources with their column layout (1-indexed).
# first/last are inclusive data-row bounds.
SOURCES = [
    {"sheet": "Adult \u2013 Original", "riddle": 2, "answer": 3, "category": 4,
     "posted": 6, "link": 7, "first": 2, "last": 21},
    {"sheet": "Adult Riddles", "riddle": 4, "answer": 5, "category": 6,
     "posted": 8, "link": 9, "first": 2, "last": 101},
]

def _is_posted(v):
    return str(v).strip().lower() in ("yes", "y", "true", "done", "1")

def find_next(wb):
    """Return dict {sheet,row,riddle,answer,seed,src} for the next unposted
    riddle in launch order, or None if the bank is exhausted.
    seed = number of already-posted adult riddles (drives weekly rotation)."""
    posted_count = 0
    nxt = None
    for src in SOURCES:
        ws = wb[src["sheet"]]
        for row in range(src["first"], src["last"] + 1):
            riddle = ws.cell(row=row, column=src["riddle"]).value
            if riddle is None or not str(riddle).strip():
                continue  # skips note/blank rows
            if _is_posted(ws.cell(row=row, column=src["posted"]).value):
                posted_count += 1
            elif nxt is None:
                answer = ws.cell(row=row, column=src["answer"]).value
                cat = ws.cell(row=row, column=src["category"]).value
                nxt = {"sheet": src["sheet"], "row": row,
                       "riddle": str(riddle).strip(), "answer": str(answer).strip(),
                       "category": str(cat).strip() if cat else "", "src": src}
    if nxt is None:
        return None
    nxt["seed"] = posted_count
    return nxt

def mark_posted(xlsx_path, sheet, row, posted_col, link_col, link):
    """Called by the uploader AFTER a successful upload. Preserves formulas."""
    wb = load_workbook(xlsx_path)  # not data_only -> keeps tracker formulas
    ws = wb[sheet]
    ws.cell(row=row, column=posted_col).value = "Yes"
    if link:
        ws.cell(row=row, column=link_col).value = link
    wb.save(xlsx_path)

def _safe(s):
    return "".join(c for c in str(s) if c.isalnum() or c in " _-").strip().replace(" ", "_")[:40]

def build_next(do_audio=True):
    wb = load_workbook(XLSX, data_only=True)
    nxt = find_next(wb)
    if nxt is None:
        print("BANK EXHAUSTED — every adult riddle is marked posted.")
        return None

    seed = nxt["seed"]
    theme = render_video.build_theme(seed)
    base = f"{nxt['sheet'].split()[0]}_s{seed:03d}_{_safe(nxt['answer'])}"
    silent = os.path.join(OUTDIR, base + ".mp4")

    print(f"NEXT  [{nxt['sheet']} r{nxt['row']}]  seed={seed}  "
          f"palette={theme['palette']['name']}")
    print(f"  Q: {nxt['riddle']}")
    print(f"  A: {nxt['answer']}")

    render_video.render(nxt["riddle"], nxt["answer"], seed, silent)
    final = silent
    if do_audio:
        r = subprocess.run([sys.executable, f"{SCRIPT_DIR}/add_audio.py", silent, "5", "15"],
                           capture_output=True, text=True)
        if r.returncode != 0:
            print("AUDIO STEP FAILED:\n", r.stderr[-800:])
        else:
            stem, ext = os.path.splitext(silent)
            cand = stem + "_ticking" + ext
            if os.path.exists(cand):
                final = cand

    meta = metadata.generate(nxt["riddle"], nxt["answer"], seed, nxt.get("category"))
    manifest = {
        "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "sheet": nxt["sheet"], "row": nxt["row"],
        "posted_col": nxt["src"]["posted"], "link_col": nxt["src"]["link"],
        "seed": seed, "riddle": nxt["riddle"], "answer": nxt["answer"],
        "category": nxt.get("category", ""),
        "palette": theme["palette"]["name"], "week": theme["week"],
        "video_path": final, "silent_path": silent,
        "title": meta["title"], "description": meta["description"],
        "tags": meta["tags"], "category_id": meta["category_id"],
    }
    with open(MANIFEST, "w") as f:
        json.dump(manifest, f, indent=2)
    print("VIDEO:", final)
    print("MANIFEST:", MANIFEST)
    return manifest

if __name__ == "__main__":
    if "--peek" in sys.argv:
        wb = load_workbook(XLSX, data_only=True)
        nxt = find_next(wb)
        print(json.dumps({k: v for k, v in (nxt or {}).items() if k != "src"},
                         indent=2, ensure_ascii=False) if nxt else "BANK EXHAUSTED")
    else:
        build_next(do_audio="--no-audio" not in sys.argv)
