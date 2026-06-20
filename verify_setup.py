#!/usr/bin/env python3
"""
Pre-flight check. Run this AFTER you've added credentials but BEFORE the first
real upload, so you confirm everything is wired correctly instead of debugging
a failed Actions run.

    python3 verify_setup.py

Checks (each independent, never aborts the rest):
  1. ffmpeg + DejaVu fonts present
  2. workbook readable + how many adult riddles remain unposted
  3. Google/YouTube auth works (prints the channel it will post to)
  4. Telegram works (sends you a real test message)
Exit code 0 = all green, 1 = something needs attention.
"""
import os, sys, shutil, subprocess

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_ROOT = os.environ.get("RIDDLE_ROOT",
                           "/sessions/magical-modest-ramanujan/mnt/Youtube")
XLSX = os.environ.get("RIDDLE_XLSX", f"{DATA_ROOT}/Riddle_Content_Bank.xlsx")
CLIENT_SECRET = os.environ.get("YT_CLIENT_SECRET", f"{SCRIPT_DIR}/client_secret.json")
TOKEN = os.environ.get("YT_TOKEN", f"{SCRIPT_DIR}/token.json")
SCOPES = ["https://www.googleapis.com/auth/youtube.upload",
          "https://www.googleapis.com/auth/youtube"]
FONT = "/usr/share/fonts/truetype/dejavu/DejaVuSerif-Bold.ttf"

ok = True
def line(good, label, detail=""):
    global ok
    ok = ok and good
    print(f"  [{'PASS' if good else 'FAIL'}] {label}" + (f"  -> {detail}" if detail else ""))

print("Riddle o'Clock — setup pre-flight\n")

# 1. tooling
print("1) Tooling")
line(shutil.which("ffmpeg") is not None, "ffmpeg installed",
     shutil.which("ffmpeg") or "not found")
line(os.path.exists(FONT), "DejaVu fonts present", FONT if os.path.exists(FONT) else "missing")

# 2. workbook
print("\n2) Content bank")
try:
    import pipeline
    from openpyxl import load_workbook
    wb = load_workbook(XLSX, data_only=True)
    nxt = pipeline.find_next(wb)
    if nxt:
        # count remaining
        remaining = 0
        for src in pipeline.SOURCES:
            ws = wb[src["sheet"]]
            for r in range(src["first"], src["last"] + 1):
                v = ws.cell(row=r, column=src["riddle"]).value
                if v and str(v).strip() and not pipeline._is_posted(
                        ws.cell(row=r, column=src["posted"]).value):
                    remaining += 1
        line(True, "workbook readable; next riddle found",
             f"{remaining} unposted left; next = '{nxt['answer']}' (seed {nxt['seed']})")
    else:
        line(False, "workbook readable but BANK EXHAUSTED", "all riddles marked posted")
except Exception as e:
    line(False, "workbook / pipeline", repr(e))

# 3. YouTube auth
print("\n3) YouTube auth")
try:
    if not os.path.exists(TOKEN):
        line(False, "token.json present", f"missing at {TOKEN} (run authorize.py)")
    else:
        from google.oauth2.credentials import Credentials
        from google.auth.transport.requests import Request
        from googleapiclient.discovery import build
        creds = Credentials.from_authorized_user_file(TOKEN, SCOPES)
        if not creds.valid and creds.refresh_token:
            creds.refresh(Request())
            open(TOKEN, "w").write(creds.to_json())
        yt = build("youtube", "v3", credentials=creds)
        ch = yt.channels().list(part="snippet", mine=True).execute()
        title = ch["items"][0]["snippet"]["title"] if ch.get("items") else "?"
        line(True, "YouTube authorized", f"posts to channel: {title}")
except Exception as e:
    line(False, "YouTube auth", repr(e))

# 4. Telegram
print("\n4) Telegram notification")
try:
    if not (os.environ.get("TELEGRAM_BOT_TOKEN") and os.environ.get("TELEGRAM_CHAT_ID")):
        line(False, "Telegram env vars set", "TELEGRAM_BOT_TOKEN / TELEGRAM_CHAT_ID missing")
    else:
        import notify
        notify.send_telegram("\u2705 Riddle o'Clock setup check — you're wired up.")
        line(True, "Telegram test message sent", "check your phone")
except Exception as e:
    line(False, "Telegram", repr(e))

print("\n" + ("ALL GREEN — safe to run the daily workflow." if ok
              else "Some checks failed — fix the FAILs above before going live."))
sys.exit(0 if ok else 1)
